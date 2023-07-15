# shurwat se for practise

# pypi wbsite hai python ki jaha se vo saarey package install karta hai
# install openpyxl
# import openpyxy
# to install --  pip install openpyxl
# koi specific version install karna to -- pip install openpyxl == 3.1.2
# to uninstall -- pip uninstall openpyxl
# agar koi naya version download karoge to vo previos version uninstall kar dega
# excel file ke aandar CRUD operations karna.
# alingment karna, merge karna sab apan openpyxl se kar saktey hai

# agar apan koi koi bhi data text file se excel me karna hai to manually bhot time lag jaaega and lenghthy kaam hai vo bahot
# islliye apan python me openpyxl use karenge 

# import openpyxl
# dheko apan excel ko workbook bhi boltey hai 
# to apan ko openpyxl se workbook import karna padega pehle 
# fir uska object banana padega neeche dheko
# fir apan ko sheet open karni padegi
# bu default jo pehli sheet hoti hai usko apan active sheet boltey hai.

#ex.
from openpyxl import Workbook,load_workbook

# wb = Workbook() # object bana liya
# active_sheet = wb.active # yaha pe sheet active ho gayi (sheet ka naam sheet hai neeche excel me dhekna)
# active_sheet.cell(row=1,column=1,value = "python") # aaise apan ko cells me data daalna hai 
# active_sheet.cell(row=2,column=2,).value = "java"# 2nd method (koi sa bhi use kar saktey ho)
# wb.save(filename="D:\oops_concept\handeling\sample_excel.xlsx") # fir yaha pe apan ne iss file ka path hi copy kar ke usko xlsx ka extension de diya to ek nayi file ban jaaegi excel ki

#### ab apan ko existing file ko open karna hai 
# to uske liye apan ko openpyxl se ek cheez or import karni padegi vo hai load_workbook

#ex.
# FILE_PATH = r"D:\oops_concept\handeling\sample_excel.xlsx" # ek constant bana liya
# wb = load_workbook(filename=FILE_PATH) # isko file path chaiye hota hai ki konsi file load karni hai vo to batao.
# activesheet = wb.active
# activesheet.cell(row=1,column=1,value="empID")
# activesheet.cell(row=1,column=2,value="empname")
# activesheet.cell(row=1,column=3,value="empsalary")
# wb.save(FILE_PATH)

#ex.2
FILE_PATH = r"D:\oops_concept\handeling\excelsample2.xlsx"
# wb = Workbook()
# sheet = wb.active
# sheet["A1"] = 87
# sheet["A2"] = "rachit"
# sheet["A3"] = 41
# sheet["A4"] = 10

# wb.save(filename=FILE_PATH)

#ex.3
# ab multiple data kaise daaley sheet me dheko
# wb = Workbook()
# sheet = wb.active
# data = (
#     (1,2,3),
#     (4,5,6),
#     (7,8,9),
#     (10,11,12),
#     (13,14,15),
# )
# for i in data: # ek ek tuple aa raha hai idhar
#     sheet.append(i) # ye list ki method nahi hai ye alag apeend hai



# wb.save(filename=FILE_PATH) # jitti baar karogey utti baar data append hotey jaaega

#ex.4
# ab apan ko read karna hai to load_workbook karna padega
# apan ko variable banana padega read karney ke liye
# dheko sirf apan read kar rahe hai to shee ko save karney ki zarurat nahi hai
# and .value se hi value niklegi dheko neeche

# wb = load_workbook(filename=FILE_PATH)
# sheet = wb.active

# x1 = sheet.cell(row=1,column=1) # iska matlab hai row 1 or column 1 me kya value hai batao
# x2 = sheet["A2"] # aaise bhi nikal saktey hai value 
# print("value of row1 and and column 1 is:-",x1.value)
# print("value of row2 and and column 1 is:-",x2.value)

#ex.5
# ab multiple cells read karna hai to
# ab yaha pe important hai for loop
# wb = load_workbook(filename=FILE_PATH)
# sheet = wb.active

# cells = sheet["A1":"B5"] # a1 se lekar b7 tak read karna hai 
# for i1 , i2 in cells: # dheko jitte column rahe utte i le lo ab c bhi chaiye to i3 le lo
#     print(f"{i1.value}--{i2.value}")

# ek or example jaise c chiye to
# wb = load_workbook(filename=FILE_PATH)
# sheet = wb.active

# cells = sheet["A1":"C5"]
# for i1,i2,i3 in cells:
#     print(f"{i1.value}--{i2.value}--{i3.value}")

#ex.6
# ITER ROWS ( ye row ke hisab se data dega )
# ab data write kar rahe hai to save karnna padega
wb = load_workbook(filename=FILE_PATH)
sheet = wb.active

# rows = (
#     (1,2,3,4),
#     (5,6,7,8),
#     (9,10,11,12),
#     (13,14,15,16),
#     (17,18,19,20),
#     (21,22,23,24),
# )
# for row in rows:
#     sheet.append(row)

# wb.save(filename=FILE_PATH)
# ab dheko apan ne data daal diya and usko commment kar diya
# kyoki apan ab read karenge data to vo everride na ho isliye
# ab read karney ke liye dheko kaise kiya
# ab iter rows use kar rahe hai data read karney ke liye 
# to iter.rows ko poora batana padega ki kitti minimum rows and column hai 

# for row in sheet.iter_rows(min_row=1, min_col=1, max_row=12, max_col=4): # isko bata re hai kitte max row and max column hai 
#     print(row)   # to yaha pe ek ek tuple aaega (1,2,3,4) aaise
#     for cell in row:  # ab vo tuple pe iterate kar rahe hai apan 
#         print(cell.value,end=" ") # ek ek cell ki value nikal ke print kar rahe hai, end bas space de raha hai
#     print() # ye print statement bas next line me dene ke liye (newline ke lie)

### ITER COLUMNS ( or ye column ke hhisab se data dega)
# iter_cols

# wb = load_workbook(filename=FILE_PATH)
# sheet = wb.active

# columns = (
#     (1,2,3,4),
#     (5,6,7,8),
#     (9,10,11,12),
#     (13,14,15,16),
#     (17,18,19,20),
#     (21,22,23,24),
#     ("yes","sir","hello","rachit"),
# )
# for col in columns:
#     sheet.append(col)

# wb.save(filename=FILE_PATH)

# for col in sheet.iter_cols(min_row = 1, min_col = 1, max_row = 19, max_col = 4):
#     for cell in col:
#         print(cell.value, end = " ")
#     print() 

#kitte sheets hai apan ke pass check karney ke liye
# wb = load_workbook(filename=FILE_PATH)
# sheet = wb.active
# print(wb.sheetnames)

#konsi sheet active hai dhekne ke liye (object return karega)
# wb =  load_workbook(filename=FILE_PATH)
# active_sheet = wb.active
# print(type(active_sheet)) #<class 'openpyxl.worksheet.worksheet.Worksheet'>

#ab koi specific sheet chaiye ho to 
# wb = load_workbook(filename=FILE_PATH)
# sheet = wb["Sheet"]
# # print(sheet.title)  #Sheet
# val = sheet.cell(row=1,column=1)
# print(val.value)


